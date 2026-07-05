#---------------------------------
# GOAM Scores & Rounds App (SPLIT VERSION)
#   - Leaderboards page
#   - Scorecards page
#---------------------------------

import io
import os
import re
import json
import hashlib
from datetime import datetime
import streamlit as st
import pandas as pd
import numpy as np
from PIL import Image, ImageDraw, ImageFont, ImageFilter
from PIL import ImageOps
from openpyxl import load_workbook

from backend.goam_loader import GOAMLoader
from backend.goam_rounds import GOAMRounds
from backend.goam_calculator import GOAMCalculator
from utils.json_utils import load_json, save_json

try:
    from rapidocr_onnxruntime import RapidOCR
except Exception:
    RapidOCR = None


# ---------------------------------------------------------
# INTERNAL STATE
# ---------------------------------------------------------
def _get_rounds_state():
    if "goam_rounds" not in st.session_state:
        st.session_state.goam_rounds = GOAMRounds()
    return st.session_state.goam_rounds


def _format_pos_change(delta):
    if delta is None:
        return "–"
    try:
        d = int(delta)
    except (TypeError, ValueError):
        return "–"

    if d > 0:
        return f"⬆️ {d}"
    if d < 0:
        return f"⬇️ {abs(d)}"
    return "➡️"


def _to_int_or_none(value):
    try:
        if value is None:
            return None
        s = str(value).strip()
        if s == "":
            return None
        return int(float(s))
    except Exception:
        return None


def _is_blank(value):
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() == ""


def _load_font(size, bold=False):
    font_candidates = [
        "arialbd.ttf" if bold else "arial.ttf",
        "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf",
    ]

    for font_name in font_candidates:
        try:
            return ImageFont.truetype(font_name, size)
        except Exception:
            continue

    return ImageFont.load_default()


def _month_key_to_mmyyyy(month_key: str) -> str:
    raw = str(month_key or "").strip().replace("’", "'")
    if not raw:
        return datetime.now().strftime("%m%Y")

    for fmt in ("%b'%y", "%b %Y", "%B %Y", "%Y-%m", "%Y/%m"):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.strftime("%m%Y")
        except Exception:
            continue

    return datetime.now().strftime("%m%Y")


def _sanitize_filename_part(value: str) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[^A-Za-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "Unknown"


def _canonicalize_players_for_fingerprint(players: list[dict]) -> list[dict]:
    canonical = []
    for p in players:
        canonical.append(
            {
                "name": str(p.get("name", "")).strip().lower(),
                "strokes": _to_int_or_none(p.get("strokes")),
                "ips": _to_int_or_none(p.get("ips")),
                "team": str(p.get("team", "")).strip().lower(),
            }
        )
    canonical.sort(key=lambda x: (x["name"], x["strokes"] if x["strokes"] is not None else -1))
    return canonical


def _scorecard_fingerprint(course_name: str, month_key: str, players: list[dict]) -> str:
    payload = {
        "course": str(course_name or "").strip().lower(),
        "month_key": str(month_key or "").strip().lower(),
        "players": _canonicalize_players_for_fingerprint(players),
    }
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _load_processed_scorecards_registry():
    data = load_json("data/processed_scorecards.json")
    if isinstance(data, dict) and isinstance(data.get("items"), list):
        return data
    return {"items": []}


def _find_duplicate_scorecard(registry: dict, fingerprint: str, image_hash: str | None):
    for item in registry.get("items", []):
        if item.get("fingerprint") == fingerprint:
            return "scorecard content", item
        if image_hash and item.get("image_hash") == image_hash:
            return "image", item
    return None, None


def _register_processed_scorecard(registry: dict, fingerprint: str, image_hash: str | None, month_key: str, course_name: str):
    item = {
        "fingerprint": fingerprint,
        "image_hash": image_hash,
        "month_key": str(month_key or "").strip(),
        "course": str(course_name or "").strip(),
        "created_at": datetime.utcnow().isoformat() + "Z",
    }
    registry.setdefault("items", []).append(item)
    save_json("data/processed_scorecards.json", registry)


def _build_course_score_from_template(course_name: str, month_key: str, player_rows: list[dict]):
    template_path = os.path.join("data", "Course-Score-Template.xlsx")
    if not os.path.exists(template_path):
        return None, "Course score template not found: data/Course-Score-Template.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_map = {
        str(h).strip().lower(): idx + 1
        for idx, h in enumerate(headers)
        if h is not None and str(h).strip() != ""
    }

    name_col = header_map.get("name", 1)
    strokes_col = header_map.get("strokes")
    ips_col = header_map.get("ips")
    team_col = header_map.get("liv")

    if strokes_col is None or ips_col is None:
        return None, "Template must include Name, Strokes, and IPS columns in row 1."

    template_row_by_name = {}
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, name_col).value or "").strip()
        if name:
            template_row_by_name[name.lower()] = r

    next_row = ws.max_row + 1
    for row in player_rows:
        name = str(row.get("name", "")).strip()
        if not name:
            continue

        r = template_row_by_name.get(name.lower())
        if r is None:
            r = next_row
            next_row += 1
            ws.cell(r, name_col).value = name

        ws.cell(r, strokes_col).value = _to_int_or_none(row.get("strokes"))
        ws.cell(r, ips_col).value = _to_int_or_none(row.get("ips"))
        if team_col is not None:
            ws.cell(r, team_col).value = str(row.get("team", "")).strip()

    safe_course = _sanitize_filename_part(course_name)
    mmYYYY = _month_key_to_mmyyyy(month_key)
    out_name = f"Course-Score-{safe_course}-{mmYYYY}.xlsx"
    out_path = os.path.join("data", out_name)
    wb.save(out_path)
    return out_path, None


def _extract_ocr_lines(uploaded_image: Image.Image):
    lines = []
    debug_text = []

    def _collect_lines(result_obj):
        out = []
        if isinstance(result_obj, list):
            for item in result_obj:
                if not isinstance(item, (list, tuple)) or len(item) < 2:
                    continue
                text = str(item[1]).strip()
                if text:
                    out.append(text)
        return out

    def _quality_score(candidate_lines):
        if not candidate_lines:
            return 0
        score = len(candidate_lines)
        for t in candidate_lines:
            lower = t.lower()
            if re.search(r"\d\s*[/|\\]\s*\d", t):
                score += 4
            if re.search(r"\d+", t):
                score += 1
            if any(k in lower for k in ("player", "marker", "total", "out", "in", "alliance")):
                score += 3
        return score

    # Preferred OCR backend: RapidOCR (fully local, no external tesseract binary required).
    if RapidOCR is not None:
        engine = RapidOCR()
        gray = uploaded_image.convert("L")
        rgb = uploaded_image.convert("RGB")
        bw = ImageOps.autocontrast(gray).point(lambda x: 255 if x > 145 else 0, mode="1").convert("RGB")
        bw_inv = ImageOps.invert(bw.convert("L")).convert("RGB")
        sharp = ImageOps.autocontrast(gray).filter(ImageFilter.SHARPEN).convert("RGB")

        candidates = [
            rgb,
            ImageOps.autocontrast(gray).convert("RGB"),
            ImageOps.equalize(gray).convert("RGB"),
            sharp,
            bw,
            bw_inv,
            rgb.rotate(-1.2, expand=True, fillcolor="white"),
            rgb.rotate(1.2, expand=True, fillcolor="white"),
        ]

        best_score = -1
        all_lines = []
        for idx, candidate in enumerate(candidates, start=1):
            arr = np.array(candidate)
            try:
                result, _ = engine(arr)
                candidate_lines = _collect_lines(result)
            except Exception as e:
                debug_text.append(f"Pass {idx}: OCR error ({e})")
                continue

            debug_text.append(f"Pass {idx}: {len(candidate_lines)} OCR lines")
            all_lines.extend(candidate_lines)

            score = _quality_score(candidate_lines)
            if score > best_score:
                best_score = score
                lines = candidate_lines

        if all_lines:
            merged = []
            seen = set()
            for t in all_lines:
                k = re.sub(r"\s+", " ", str(t).strip().lower())
                if not k or k in seen:
                    continue
                seen.add(k)
                merged.append(str(t).strip())

            merged_score = _quality_score(merged)
            if merged_score >= max(best_score, 0):
                lines = merged

    return lines, "\n".join(debug_text)


def _parse_ocr_scorecard_lines(lines):
    parsed = []
    seen_names = set()

    for raw in lines:
        line = re.sub(r"\s+", " ", str(raw or "")).strip()
        if not line:
            continue

        nums = list(re.finditer(r"\d+", line))
        if len(nums) < 2:
            continue

        # Assume the last two numbers are Strokes and IPS.
        strokes_match = nums[-2]
        ips_match = nums[-1]

        try:
            strokes = int(strokes_match.group())
            ips = int(ips_match.group())
        except Exception:
            continue

        if not (40 <= strokes <= 160 and 0 <= ips <= 60):
            continue

        name_part = line[:strokes_match.start()].strip(" -|:;,.\t")
        name_part = re.sub(r"[^A-Za-z '\-]", "", name_part).strip()
        if len(name_part) < 3:
            continue

        # Title-case improves readability while still editable by the user.
        name = " ".join(token.capitalize() for token in name_part.split())
        name_key = name.lower()
        if name_key in seen_names:
            continue
        seen_names.add(name_key)

        parsed.append(
            {
                "Name": name,
                "Strokes": strokes,
                "IPS": ips,
                "LIV": "",
                "OCR Line": line,
            }
        )

    parsed = sorted(parsed, key=lambda r: int(r.get("IPS", 0)), reverse=True)
    return pd.DataFrame(parsed)


def _clean_ocr_name(text):
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    value = re.sub(r"\b(sap\s*id|name|marker\s*[a-d]|player\s*[a-d])\b", "", value, flags=re.IGNORECASE)
    value = re.sub(r"[^A-Za-z '\-]", "", value).strip(" -|:;,.\t")
    if len(value) < 3:
        return ""
    return " ".join(token.capitalize() for token in value.split())


def _extract_slot_names(lines):
    slots = {"A": "", "B": "", "C": "", "D": ""}
    patterns = [
        re.compile(r"\bplayer\s*([abcd])\b\s*[:\-\s]*(.+)$", flags=re.IGNORECASE),
        re.compile(r"\bmarker\s*([abcd])\b\s*[:\-\s]*(.+)$", flags=re.IGNORECASE),
        re.compile(r"\bplay\s*([abcd])\b\s*[:\-\s]*(.+)$", flags=re.IGNORECASE),
    ]

    for raw in lines:
        line = re.sub(r"\s+", " ", str(raw or "")).strip()
        if not line:
            continue
        for pat in patterns:
            m = pat.search(line)
            if not m:
                continue
            slot = m.group(1).upper()
            tail = m.group(2)
            name = _clean_ocr_name(tail)
            if name and not slots[slot]:
                slots[slot] = name

    return slots


def _extract_score_ips_pairs(text):
    pairs = []
    line = str(text or "")

    for m in re.finditer(r"(\d{1,3})\s*[/|\\]\s*(\d{1,2})", line):
        strokes = _to_int_or_none(m.group(1))
        ips = _to_int_or_none(m.group(2))
        if strokes is None or ips is None:
            continue
        if 40 <= strokes <= 160 and 0 <= ips <= 60:
            pairs.append((strokes, ips))

    if len(pairs) >= 4:
        return pairs[:4]

    nums = [_to_int_or_none(x) for x in re.findall(r"\d+", line)]
    nums = [n for n in nums if n is not None]
    if len(nums) >= 8:
        candidate = []
        for i in range(0, min(8, len(nums)), 2):
            strokes = nums[i]
            ips = nums[i + 1]
            if 40 <= strokes <= 160 and 0 <= ips <= 60:
                candidate.append((strokes, ips))
        if len(candidate) >= 4:
            return candidate[:4]

    return pairs


def _parse_scorecard_layout_a_to_d(lines):
    slots = _extract_slot_names(lines)
    ordered_slots = ["A", "B", "C", "D"]

    # Parse the total row where score/result pairs are typically written.
    total_line = ""
    total_numbers = []
    best_pair_line = ""
    best_pairs = []
    for raw in lines:
        line = re.sub(r"\s+", " ", str(raw or "")).strip()
        if not line:
            continue

        pairs = _extract_score_ips_pairs(line)
        lower = line.lower()
        weight = 2 if any(k in lower for k in ("total", "out", "in", "alliance")) else 0
        if len(pairs) + weight > len(best_pairs):
            best_pairs = pairs
            best_pair_line = line

        if "total" in lower or "out" in lower or "in" in lower:
            nums = [int(x) for x in re.findall(r"\d+", line)]
            if len(nums) > len(total_numbers):
                total_numbers = nums
                total_line = line

    if not any(slots.values()):
        for slot in ordered_slots:
            slots[slot] = f"Player {slot}"

    if len(best_pairs) < 4 and len(total_numbers) < 4:
        return pd.DataFrame()

    values = total_numbers[:]
    if values and values[0] in {36, 72}:
        values = values[1:]

    rows = []
    if len(best_pairs) >= 4:
        for i, slot in enumerate(ordered_slots):
            name = slots.get(slot, "")
            if not name:
                continue
            strokes, ips = best_pairs[i]
            rows.append(
                {
                    "Name": name,
                    "Strokes": strokes,
                    "IPS": ips,
                    "LIV": "",
                    "OCR Line": best_pair_line,
                }
            )
    elif len(values) >= 8:
        # Assume score/result pairs for each player.
        for i, slot in enumerate(ordered_slots):
            name = slots.get(slot, "")
            if not name:
                continue
            strokes = int(values[i * 2])
            ips = int(values[i * 2 + 1])
            if not (40 <= strokes <= 160 and 0 <= ips <= 60):
                continue
            rows.append(
                {
                    "Name": name,
                    "Strokes": strokes,
                    "IPS": ips,
                    "LIV": "",
                    "OCR Line": total_line,
                }
            )
    else:
        # Fallback: use first four numbers as strokes only; leave IPS empty for manual edit.
        for i, slot in enumerate(ordered_slots):
            if i >= len(values):
                break
            name = slots.get(slot, "")
            if not name:
                continue
            strokes = int(values[i])
            if not (40 <= strokes <= 160):
                continue
            rows.append(
                {
                    "Name": name,
                    "Strokes": strokes,
                    "IPS": "",
                    "LIV": "",
                    "OCR Line": total_line,
                }
            )

    out = pd.DataFrame(rows)
    if not out.empty and "IPS" in out.columns:
        try:
            numeric_ips = pd.to_numeric(out["IPS"], errors="coerce")
            out = out.assign(_ips_sort=numeric_ips.fillna(-1)).sort_values("_ips_sort", ascending=False).drop(columns=["_ips_sort"])
        except Exception:
            pass
    return out


def _compute_nett(player):
    strokes = _to_int_or_none(player.get("strokes"))
    handicap = _to_int_or_none(player.get("handicap"))
    if strokes is None or handicap is None:
        return None
    return strokes - handicap


def _build_scoreboard_image(month_key, month_data, best_gross_override=None, best_nett_override=None):
    players = month_data.get("players", [])
    course_name = str(month_data.get("course", "")).strip() or "Unknown Course"
    title = f"GOAM at {course_name} — {month_key}"

    valid_players = [p for p in players if str(p.get("name", "")).strip()]
    sorted_players = sorted(valid_players, key=lambda p: (_to_int_or_none(p.get("ips")) or -9999), reverse=True)
    top3 = sorted_players[:3]
    full_field = [
        [
            str(p.get("name", "")).strip(),
            _to_int_or_none(p.get("strokes")) or "",
            _to_int_or_none(p.get("ips")) or "",
        ]
        for p in sorted_players
    ]

    best_gross_name = month_data.get("best_gross")
    best_gross_score = None
    if best_gross_name:
        best_gross_player = next((p for p in valid_players if p.get("name") == best_gross_name), None)
        if best_gross_player:
            best_gross_score = _to_int_or_none(best_gross_player.get("strokes"))

    if best_gross_override:
        override_name = str(best_gross_override.get("name", "")).strip()
        override_score = _to_int_or_none(best_gross_override.get("score"))
        if override_name:
            best_gross_name = override_name
        if override_score is not None:
            best_gross_score = override_score

    best_nett_name = month_data.get("best_nett")
    best_nett_score = None
    if best_nett_name:
        best_nett_player = next((p for p in valid_players if p.get("name") == best_nett_name), None)
        if best_nett_player:
            best_nett_score = _compute_nett(best_nett_player)

    if best_nett_override:
        override_name = str(best_nett_override.get("name", "")).strip()
        override_score = _to_int_or_none(best_nett_override.get("score"))
        if override_name:
            best_nett_name = override_name
        if override_score is not None:
            best_nett_score = override_score

    def award_lines(field_name):
        lines = []
        for player in sorted_players:
            value = player.get(field_name)
            if _is_blank(value):
                continue
            hole = _to_int_or_none(value)
            if hole is None:
                lines.append(str(player.get("name", "")))
            else:
                lines.append(f"{player.get('name', '')} (H{hole})")
        return ", ".join(lines) if lines else "-"

    def merge_awards(*keys):
        merged = []
        for key in keys:
            values = award_lines(key)
            if values == "-":
                continue
            merged.extend([v.strip() for v in values.split(",") if v.strip()])
        return ", ".join(merged) if merged else "-"

    np_line = f"NP: {merge_awards('np1', 'np2')}"
    ld_line = f"LD: {merge_awards('ld1', 'ld2')}"

    liv_totals = month_data.get("liv_totals") or {}
    if not liv_totals:
        grouped = {}
        for player in valid_players:
            team = str(player.get("team", "")).strip()
            ips = _to_int_or_none(player.get("ips"))
            if not team or ips is None:
                continue
            grouped.setdefault(team, []).append(ips)
        liv_totals = {team: sum(sorted(values, reverse=True)[:3]) for team, values in grouped.items()}

    pools = []
    for player in valid_players:
        pool_bet = player.get("pool_bet")
        if _is_blank(pool_bet):
            continue
        pools.append((str(player.get("name", "")), str(pool_bet)))

    fines_total = _to_int_or_none(month_data.get("fines_total"))
    if fines_total is None:
        fines_total = sum(_to_int_or_none(player.get("fines")) or 0 for player in valid_players)

    width, height = 1200, 1800
    img = Image.new("RGB", (width, height), (244, 225, 185))
    draw = ImageDraw.Draw(img)

    title_font = _load_font(56, bold=True)
    section_font = _load_font(42, bold=True)
    body_font = _load_font(34)
    body_bold_font = _load_font(35, bold=True)
    small_font = _load_font(30)

    # Parchment gradient + subtle texture.
    for y in range(height):
        shade = int(242 - (y / float(height)) * 26)
        draw.line([(0, y), (width, y)], fill=(shade, shade - 16, shade - 48))

    for y in range(40, height, 22):
        for x in range(30, width, 24):
            tone = 186 + ((x + y) % 18)
            draw.point((x, y), fill=(tone, tone - 18, tone - 40))

    red = (148, 17, 17)
    blue = (22, 52, 103)
    ink = (60, 39, 24)

    # Border and center guide.
    draw.rounded_rectangle((14, 14, width - 14, height - 14), radius=20, outline=(158, 103, 60), width=6)
    divider_x = 585
    draw.line((divider_x, 185, divider_x, height - 180), fill=(170, 116, 74), width=3)

    title_bbox = draw.textbbox((0, 0), title, font=title_font)
    title_w = title_bbox[2] - title_bbox[0]
    title_x = max(42, (width - title_w) // 2)
    draw.text((title_x, 58), title, fill=red, font=title_font)
    draw.line((95, 135, width - 95, 135), fill=red, width=4)

    # Decorative icon sketches to mimic the reference poster style.
    draw.ellipse((20, 255, 160, 325), fill=(98, 145, 54), outline=blue, width=3)
    draw.rectangle((98, 170, 104, 265), fill=ink)
    draw.polygon([(104, 174), (164, 194), (104, 220)], fill=(184, 30, 24), outline=ink)

    cart_x, cart_y = 820, 182
    draw.rectangle((cart_x, cart_y + 32, cart_x + 95, cart_y + 72), outline=blue, width=4)
    draw.rectangle((cart_x + 55, cart_y + 5, cart_x + 110, cart_y + 45), outline=blue, width=4)
    draw.ellipse((cart_x + 8, cart_y + 66, cart_x + 34, cart_y + 94), outline=blue, width=4)
    draw.ellipse((cart_x + 72, cart_y + 66, cart_x + 98, cart_y + 94), outline=blue, width=4)
    draw.ellipse((962, 198, 1025, 262), outline=blue, width=4)
    draw.rectangle((991, 262, 997, 286), fill=blue)

    draw.ellipse((1010, 760, 1085, 835), outline=blue, width=4)
    draw.polygon([(1047, 835), (1025, 878), (1068, 878)], fill=blue)

    draw.line((66, 1360, 96, 1540), fill=blue, width=6)
    draw.line((92, 1364, 136, 1536), fill=blue, width=6)
    draw.ellipse((52, 1344, 100, 1382), outline=blue, width=4)
    draw.ellipse((82, 1344, 132, 1382), outline=blue, width=4)

    draw.rectangle((1000, 1500, 1110, 1670), outline=blue, width=4)
    draw.rectangle((1110, 1540, 1136, 1652), outline=blue, width=4)
    draw.arc((1004, 1468, 1110, 1548), 180, 360, fill=blue, width=4)
    draw.ellipse((1010, 1464, 1118, 1540), fill=(236, 232, 222), outline=blue, width=2)

    def section_header(x, y, text):
        draw.text((x, y), text, fill=red, font=section_font)
        draw.line((x, y + 46, x + 300, y + 46), fill=(176, 112, 66), width=2)

    # Left column.
    section_header(96, 246, "TOP 3 (IPS):")
    y = 318
    for idx, player in enumerate(top3, 1):
        name = str(player.get("name", "")).strip()
        ips = _to_int_or_none(player.get("ips"))
        draw.text((112, y), f"{idx}. {name} {ips if ips is not None else ''}", fill=blue, font=body_bold_font)
        y += 46

    section_header(22, 470, "* BEST GROSS:")
    draw.text(
        (66, 530),
        f"{best_gross_name or '-'} {best_gross_score if best_gross_score is not None else ''}",
        fill=blue,
        font=body_bold_font,
    )

    section_header(22, 610, "* BEST NETT:")
    draw.text(
        (66, 670),
        f"{best_nett_name or '-'} {best_nett_score if best_nett_score is not None else ''}",
        fill=blue,
        font=body_bold_font,
    )

    section_header(22, 748, "* NP & LD:")
    draw.text((66, 807), np_line, fill=blue, font=small_font)
    draw.text((66, 852), ld_line, fill=blue, font=small_font)

    right_header_x = 595
    right_name_x = 600
    right_strokes_x = 905
    right_ips_x = 995

    section_header(right_header_x, 246, "FULL FIELD:")
    y = 318
    for name, strokes, ips in full_field:
        draw.text((right_name_x, y), str(name), fill=blue, font=body_font)
        draw.text((right_strokes_x, y), str(strokes), fill=ink, font=body_bold_font)
        draw.text((right_ips_x, y), str(ips), fill=blue, font=body_bold_font)
        y += 40

    section_header(22, 940, "* LIV TEAMS:")
    y = 1000
    ranked_teams = sorted(liv_totals.items(), key=lambda item: item[1], reverse=True)
    for team, total in ranked_teams:
        draw.text((66, y), f"{team} {total}", fill=blue, font=body_font)
        y += 42

    y = max(y + 22, 1668)
    draw.line((28, y - 14, width - 28, y - 14), fill=(176, 112, 66), width=2)
    pools_text = "  ".join(f"{name} {amount}" for name, amount in pools) if pools else "None"
    draw.text((30, y), f"POOLS: {pools_text}", fill=red, font=body_bold_font)
    draw.text((900, y), "FINES:", fill=red, font=body_bold_font)
    draw.text((905, y + 42), f"R{fines_total} Each", fill=blue, font=body_bold_font)

    image_buffer = io.BytesIO()
    img.save(image_buffer, format="PNG")
    image_buffer.seek(0)
    return image_buffer.getvalue()


def _generated_scorecard_to_rows():
    generated_scorecard = load_json("data/generated_scorecard.json")
    if not isinstance(generated_scorecard, dict):
        return pd.DataFrame(), "Generated scorecard file missing or invalid."

    month_key = str(generated_scorecard.get("month_key", "")).strip()
    course_name = str(generated_scorecard.get("course", "")).strip()
    rows = generated_scorecard.get("scorecard", [])

    if not rows:
        return pd.DataFrame(), "No generated scorecard rows found."
    if not month_key:
        return pd.DataFrame(), "Generated scorecard is missing month_key."
    if not course_name:
        return pd.DataFrame(), "Generated scorecard is missing course name."

    out = []
    for idx, row in enumerate(rows, 1):
        name = str(row.get("Name", "")).strip()
        if not name:
            continue

        strokes = _to_int_or_none(row.get("Strokes"))
        ips = _to_int_or_none(row.get("IPS"))
        if strokes is None or ips is None:
            continue

        out.append(
            {
                "Name": name,
                "Strokes": strokes,
                "IPS": ips,
                "Course": course_name,
                "Month": month_key,
                "Team": str(row.get("LIV", "")).strip(),
            }
        )

    if not out:
        return pd.DataFrame(), "Generated scorecard has no valid rows with numeric Strokes and IPS."

    return pd.DataFrame(out), None


# ---------------------------------------------------------
# LOAD + PREPARE DATA (shared by both pages)
# ---------------------------------------------------------
def _load_scores():
    rounds = _get_rounds_state()

    try:
        goam_scores = GOAMLoader.load_json_scores("data/goam_scores.json")
        season_rounds = GOAMCalculator.build_from_json(goam_scores)
        # --- ADD ROUND NUMBERS ---
        season_rounds = season_rounds.sort_values(["Course"]).reset_index(drop=True)
        season_rounds["Round"] = season_rounds.groupby(["Course"]).ngroup() + 1


        if not season_rounds.empty:
            # Reset stored rounds
            rounds.rounds = []
            
            for rnd in sorted(season_rounds["Round"].unique()):
                df_round = season_rounds[season_rounds["Round"] == rnd]
                leaderboard = GOAMCalculator.build_ips_leaderboard(df_round)
                rounds.update_position_history(leaderboard)
                rounds.rounds.append(df_round)
        else:
            return None, None, "No GOAM scores found. Load data via Data Manager."

    except Exception as e:
        return None, None, f"Error loading GOAM scores: {e}"

    all_rounds_df = rounds.get_all_rounds()
    if all_rounds_df.empty:
        return None, None, "No rounds available."

    return rounds, all_rounds_df, None


# ---------------------------------------------------------
# PAGE 1 — LEADERBOARDS
# ---------------------------------------------------------
def show_leaderboards():
    st.header("📘 GOAM Scores & Rounds — Leaderboards")

    rounds, all_rounds_df, error = _load_scores()
 
    if error:
        st.error(error)
        return

    include_generated = st.checkbox(
        "Include generated scorecard (without publishing)",
        value=False,
        help="Temporarily adds data/generated_scorecard.json to leaderboard calculations.",
    )

    if include_generated:
        generated_df, generated_error = _generated_scorecard_to_rows()
        if generated_error:
            st.warning(generated_error)
        elif not generated_df.empty:
            all_rounds_df = pd.concat([all_rounds_df, generated_df], ignore_index=True)
            st.caption(
                f"Included generated scorecard: {generated_df['Course'].iloc[0]} "
                f"({generated_df['Month'].iloc[0]}) with {len(generated_df)} players."
            )

    # Course selection
    st.subheader("🎯 Select courses to include in leaderboards")

    all_courses = GOAMCalculator.list_courses(all_rounds_df)
    active_courses = GOAMCalculator.get_active_courses(all_rounds_df)

    selected_courses = st.multiselect(
        "Only include these courses:",
        all_courses,
        default=active_courses
    )

    filtered_df = all_rounds_df[all_rounds_df["Course"].isin(selected_courses)]

    # Leaderboard calculations
    ips_table = GOAMCalculator.build_ips_leaderboard(filtered_df)
    strokes_table = GOAMCalculator.build_strokes_leaderboard(filtered_df)
    liv_table = GOAMCalculator.build_liv_leaderboard(filtered_df)

    if ips_table.empty:
        st.info("No IPS data available for selected courses.")
        return

    ips_table.rename(columns={c: c.strip() for c in ips_table.columns}, inplace=True)

    if "Position" not in ips_table.columns:
        if "IPS" in ips_table.columns:
            ips_table["Position"] = (
                ips_table["IPS"]
                .rank(ascending=False, method="min")
                .astype(int)
            )
        else:
            st.error("IPS column missing from IPS leaderboard.")
            return

    # 🔥 HERE: calculate position change inside GOAMRounds
    rounds.update_position_history(ips_table)

    ips_table = ips_table.copy()

    # Position Movement column
    movement = GOAMCalculator.calculate_position_movement(filtered_df)
    if movement and "Name" in ips_table.columns:
        ips_table.insert(2, "Pos. Mov.", ips_table["Name"].map(movement).fillna("–"))

    # Drop internal Position column before display
    display_table = ips_table.drop(columns=["Position"])

    # Leaderboard selector
    st.subheader("🏆 Leaderboards")

    leaderboard_choice = st.selectbox(
        "Select leaderboard:",
        ["IPS", "Strokes", "LIV"],
        index=0
    )

    if leaderboard_choice == "IPS":
        st.subheader("🏆 IPS Leaderboard (Best 6 + Course Breakdown)")
        st.dataframe(display_table, hide_index=True, use_container_width=True)
    elif leaderboard_choice == "Strokes":
        st.subheader("⛳ Strokes Leaderboard (Best 6 Over Par)")
        st.dataframe(strokes_table, hide_index=True, use_container_width=True)
    elif leaderboard_choice == "LIV":
        st.subheader("🏁 LIV Team Leaderboard (Top 3 IPS per Course)")
        st.dataframe(liv_table, hide_index=True, use_container_width=True)


def show_scoreboard_poster():
    st.header("🖼️ GOAM Poster Scoreboard")

    goam_scores = load_json("data/goam_scores.json")
    if not isinstance(goam_scores, dict) or not goam_scores:
        st.info("No GOAM scores found to generate a poster.")
        return

    month_options = list(goam_scores.keys())
    scoreboard_month = st.selectbox(
        "Select month for poster image",
        ["None"] + month_options,
        key="poster_month_select",
    )

    if scoreboard_month != "None" and scoreboard_month in goam_scores:
        scoreboard_data = goam_scores[scoreboard_month]
        best_gross_override_name = st.text_input(
            "Best gross override name",
            value=str(scoreboard_data.get("best_gross") or ""),
            key="poster_best_gross_name",
        )
        default_best_gross_score = ""
        best_gross_player = next(
            (p for p in scoreboard_data.get("players", []) if p.get("name") == scoreboard_data.get("best_gross")),
            None,
        )
        if best_gross_player and best_gross_player.get("strokes") is not None:
            default_best_gross_score = str(int(float(best_gross_player.get("strokes"))))
        best_gross_override_score = st.text_input(
            "Best gross override score",
            value=default_best_gross_score,
            key="poster_best_gross_score",
        )

        best_nett_override_name = st.text_input(
            "Best nett override name",
            value=str(scoreboard_data.get("best_nett") or ""),
            key="poster_best_nett_name",
        )
        default_best_nett_score = ""
        best_nett_player = next(
            (p for p in scoreboard_data.get("players", []) if p.get("name") == scoreboard_data.get("best_nett")),
            None,
        )
        computed_nett = _compute_nett(best_nett_player or {})
        if computed_nett is not None:
            default_best_nett_score = str(computed_nett)
        best_nett_override_score = st.text_input(
            "Best nett override score",
            value=default_best_nett_score,
            key="poster_best_nett_score",
        )

        if st.button("Generate Poster Scoreboard"):
            image_bytes = _build_scoreboard_image(
                scoreboard_month,
                scoreboard_data,
                best_gross_override={
                    "name": best_gross_override_name,
                    "score": best_gross_override_score,
                },
                best_nett_override={
                    "name": best_nett_override_name,
                    "score": best_nett_override_score,
                },
            )
            st.session_state["poster_image_bytes"] = image_bytes
            st.session_state["poster_image_filename"] = (
                f"GOAM_{scoreboard_month.replace("'", "").replace(' ', '_')}_Poster.png"
            )

    image_bytes = st.session_state.get("poster_image_bytes")
    if image_bytes:
        st.image(image_bytes, use_container_width=True)
        st.download_button(
            label="Download poster image",
            data=image_bytes,
            file_name=st.session_state.get("poster_image_filename", "goam_poster.png"),
            mime="image/png",
            use_container_width=True,
        )


# ---------------------------------------------------------
# PAGE 2 — SCORECARDS
# ---------------------------------------------------------
def show_scorecards():
    st.header("📂 GOAM Scorecards")

    goam_scores = load_json("data/goam_scores.json")
    rounds, all_rounds_df, error = _load_scores()
    if error:
        st.error(error)
        return

    # Build course sheets
    course_sheets = GOAMCalculator.split_by_course(all_rounds_df)

    st.subheader("📄 View Score Cards")

    options = ["None"] + list(course_sheets.keys())
    choice = st.selectbox("Select Course", options)

    if choice in course_sheets:
        st.dataframe(course_sheets[choice], hide_index=True, use_container_width=True)

    st.subheader("🧾 Scorecard from 4-Ball Generator")
    generated_scorecard = load_json("data/generated_scorecard.json")

    if isinstance(generated_scorecard, dict):
        generated_rows = generated_scorecard.get("scorecard", [])
        generated_month_key = str(generated_scorecard.get("month_key", "")).strip()
        generated_course_name = str(generated_scorecard.get("course", "")).strip()
    else:
        generated_rows = []
        generated_month_key = ""
        generated_course_name = ""

    if generated_rows:
        generated_df = pd.DataFrame(generated_rows)

        month_key = st.text_input(
            "Month key for this generated scorecard",
            value=generated_month_key,
            key="generated_scorecard_month_key",
            help="Example: Jul'26",
        )
        course_name = st.text_input(
            "Course name for this generated scorecard",
            value=generated_course_name,
            key="generated_scorecard_course_name",
        )

        edited_generated_df = st.data_editor(
            generated_df,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            disabled=["Fourball"],
            key="generated_scorecard_editor",
        )

        generated_df = edited_generated_df

        if st.button("Save generated scorecard edits"):
            cleaned_rows = edited_generated_df.fillna("").to_dict(orient="records")
            save_json(
                "data/generated_scorecard.json",
                {
                    "month_key": month_key,
                    "course": course_name,
                    "scorecard": cleaned_rows,
                },
            )
            st.success("Generated scorecard updated.")

        if st.button("Publish generated scorecard to GOAM scores"):
            month_key_clean = month_key.strip()
            course_name_clean = course_name.strip()

            if not month_key_clean:
                st.error("Month key is required before publishing.")
            elif not course_name_clean:
                st.error("Course name is required before publishing.")
            else:
                records = edited_generated_df.fillna("").to_dict(orient="records")
                invalid_rows = []
                players = []

                optional_map = {
                    "Handicap": "handicap",
                    "NP1": "np1",
                    "NP2": "np2",
                    "LD1": "ld1",
                    "LD2": "ld2",
                    "BG": "bg",
                    "BN": "bn",
                    "Pool Bet": "pool_bet",
                    "Pool Payouts": "pool_payouts",
                    "Fines": "fines",
                }

                for idx, row in enumerate(records, 1):
                    name = str(row.get("Name", "")).strip()
                    if not name:
                        continue

                    strokes = _to_int_or_none(row.get("Strokes"))
                    ips = _to_int_or_none(row.get("IPS"))

                    if strokes is None or ips is None:
                        invalid_rows.append(idx)
                        continue

                    player = {
                        "name": name,
                        "strokes": strokes,
                        "ips": ips,
                        "team": str(row.get("LIV", "")).strip(),
                    }

                    for source_col, target_key in optional_map.items():
                        value = _to_int_or_none(row.get(source_col))
                        if value is not None:
                            player[target_key] = value

                    players.append(player)

                if invalid_rows:
                    st.error(
                        "Strokes and IPS must be numeric for all players before publish. "
                        f"Invalid rows: {invalid_rows}"
                    )
                elif not players:
                    st.error("No valid player rows found to publish.")
                else:
                    goam_scores = load_json("data/goam_scores.json")
                    if not isinstance(goam_scores, dict):
                        goam_scores = {}

                    goam_scores[month_key_clean] = {
                        "course": course_name_clean,
                        "players": players,
                    }

                    save_json("data/goam_scores.json", goam_scores)
                    save_json(
                        "data/generated_scorecard.json",
                        {
                            "month_key": month_key_clean,
                            "course": course_name_clean,
                            "scorecard": records,
                        },
                    )

                    course_score_path, course_score_error = _build_course_score_from_template(
                        course_name=course_name_clean,
                        month_key=month_key_clean,
                        player_rows=players,
                    )

                    st.success(
                        f"Published generated scorecard to data/goam_scores.json under {month_key_clean}."
                    )
                    if course_score_error:
                        st.warning(course_score_error)
                    elif course_score_path:
                        st.success(f"Created course score file: {course_score_path}")
                    st.rerun()
    else:
        st.info("No generated scorecard found yet. Create one in 4-Ball Generation first.")

    include_generated_export = st.checkbox(
        "Include generated scorecard in workbook export (without publishing)",
        value=False,
        help="When enabled, export calculations include data/generated_scorecard.json.",
        key="include_generated_export_toggle",
    )

    export_df = all_rounds_df.copy()
    generated_export_df = pd.DataFrame()

    if include_generated_export:
        generated_export_df, generated_export_error = _generated_scorecard_to_rows()
        if generated_export_error:
            st.warning(generated_export_error)
        elif not generated_export_df.empty:
            export_df = pd.concat([export_df, generated_export_df], ignore_index=True)
            st.caption(
                f"Workbook export includes generated scorecard: {generated_export_df['Course'].iloc[0]} "
                f"({generated_export_df['Month'].iloc[0]}) with {len(generated_export_df)} players."
            )

    # Export workbook
    st.subheader("💾 Export updated GOAM workbook")

    ips_table = GOAMCalculator.build_ips_leaderboard(export_df)
    strokes_table = GOAMCalculator.build_strokes_leaderboard(export_df)
    liv_table = GOAMCalculator.build_liv_leaderboard(export_df)

    export_course_sheets = GOAMCalculator.split_by_course(export_df)

    output_file = GOAMCalculator.generate_output_filename()
    os.makedirs("data", exist_ok=True)
    output_path = os.path.join("data", output_file)

    with pd.ExcelWriter(output_path) as writer:
        ips_table.to_excel(writer, sheet_name="IPS", index=False)
        strokes_table.to_excel(writer, sheet_name="Strokes", index=False)
        liv_table.to_excel(writer, sheet_name="LIV", index=False)

        if generated_rows:
            generated_df.to_excel(writer, sheet_name="GeneratedScorecard", index=False)

        for course, df in export_course_sheets.items():
            df.to_excel(writer, sheet_name=course, index=False)

    with open(output_path, "rb") as f:
        st.download_button(
            label=f"Download {output_file}",
            data=f.read(),
            file_name=output_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


def show_scorecard_ocr_reader():
    st.header("📷 Scorecard OCR Reader")
    st.caption("Upload a physical scorecard image, extract rows with OCR, review them, and publish to GOAM scores.")

    uploaded = st.file_uploader(
        "Upload scorecard image",
        type=["png", "jpg", "jpeg", "webp"],
        key="scorecard_ocr_upload",
    )

    if uploaded is None:
        st.info("Upload a scorecard image to begin OCR extraction.")
        return

    try:
        image = Image.open(uploaded).convert("RGB")
        image_bytes = uploaded.getvalue()
        image_hash = hashlib.sha256(image_bytes).hexdigest()
        st.session_state["ocr_uploaded_image_hash"] = image_hash
    except Exception as e:
        st.error(f"Could not read image: {e}")
        return

    st.image(image, caption="Uploaded scorecard", use_container_width=True)

    if RapidOCR is None:
        st.warning(
            "OCR engine is not available. Install dependency 'rapidocr-onnxruntime' to enable extraction."
        )
        return

    if st.button("Run OCR Extraction", use_container_width=True, key="run_scorecard_ocr"):
        with st.spinner("Running OCR extraction..."):
            try:
                lines, debug_text = _extract_ocr_lines(image)
                parsed_df = _parse_ocr_scorecard_lines(lines)
                layout_df = _parse_scorecard_layout_a_to_d(lines)

                if parsed_df.empty and not layout_df.empty:
                    parsed_df = layout_df
                elif not parsed_df.empty and not layout_df.empty:
                    merged = pd.concat([parsed_df, layout_df], ignore_index=True)
                    merged = merged.drop_duplicates(subset=["Name"], keep="first")
                    parsed_df = merged

                st.session_state["ocr_scorecard_lines"] = lines
                st.session_state["ocr_scorecard_debug"] = debug_text
                st.session_state["ocr_scorecard_df"] = parsed_df
                st.session_state["ocr_scorecard_last_run"] = {
                    "line_count": len(lines),
                    "row_count": 0 if parsed_df is None else int(len(parsed_df.index)),
                }
            except Exception as e:
                st.session_state["ocr_scorecard_lines"] = []
                st.session_state["ocr_scorecard_debug"] = f"OCR extraction failed: {e}"
                st.session_state["ocr_scorecard_df"] = pd.DataFrame()
                st.session_state["ocr_scorecard_last_run"] = {
                    "line_count": 0,
                    "row_count": 0,
                }
                st.error(f"OCR extraction failed: {e}")

    last_run = st.session_state.get("ocr_scorecard_last_run")
    if last_run:
        line_count = int(last_run.get("line_count", 0))
        row_count = int(last_run.get("row_count", 0))
        if row_count > 0:
            st.success(f"OCR complete: {line_count} text lines detected, {row_count} score rows parsed.")
        else:
            st.warning(
                f"OCR complete: {line_count} text lines detected, but no valid score rows were parsed. "
                "Use a clearer image or edit rows manually."
            )

    debug_text = st.session_state.get("ocr_scorecard_debug", "")
    if debug_text:
        st.caption(debug_text)

    lines = st.session_state.get("ocr_scorecard_lines", [])
    if lines:
        with st.expander("Raw OCR text", expanded=False):
            st.text("\n".join(lines))

    parsed_df = st.session_state.get("ocr_scorecard_df")
    if parsed_df is None:
        st.info("Click 'Run OCR Extraction' to parse players, strokes, and IPS.")
        return

    if parsed_df.empty:
        st.warning("OCR ran, but no valid score rows were detected. Try a clearer photo or edit manually after extraction.")
        return

    st.subheader("Review parsed rows")
    editable_df = st.data_editor(
        parsed_df,
        hide_index=True,
        use_container_width=True,
        num_rows="dynamic",
        key="ocr_scorecard_editor",
    )

    col1, col2 = st.columns(2)
    with col1:
        month_key = st.text_input("Month key", value="", help="Example: Jul'26", key="ocr_month_key")
    with col2:
        course_name = st.text_input("Course name", value="", key="ocr_course_name")

    overwrite = st.checkbox("Overwrite existing month if it already exists", value=False, key="ocr_overwrite_month")

    if st.button("Publish OCR scorecard to GOAM scores", use_container_width=True, key="publish_ocr_scorecard"):
        month_key_clean = str(month_key).strip()
        course_name_clean = str(course_name).strip()

        if not month_key_clean:
            st.error("Month key is required.")
            return
        if not course_name_clean:
            st.error("Course name is required.")
            return

        goam_scores = load_json("data/goam_scores.json")
        if not isinstance(goam_scores, dict):
            goam_scores = {}

        if month_key_clean in goam_scores and not overwrite:
            st.error("Month already exists in GOAM scores. Enable overwrite to replace it.")
            return

        rows = editable_df.fillna("").to_dict(orient="records")
        players = []
        invalid_rows = []

        for idx, row in enumerate(rows, 1):
            name = str(row.get("Name", "")).strip()
            if not name:
                continue

            strokes = _to_int_or_none(row.get("Strokes"))
            ips = _to_int_or_none(row.get("IPS"))
            if strokes is None or ips is None:
                invalid_rows.append(idx)
                continue

            players.append(
                {
                    "name": name,
                    "strokes": strokes,
                    "ips": ips,
                    "team": str(row.get("LIV", "")).strip(),
                }
            )

        if invalid_rows:
            st.error(f"Invalid numeric rows found: {invalid_rows}. Fix them before publishing.")
            return
        if not players:
            st.error("No valid player rows to publish.")
            return

        fingerprint = _scorecard_fingerprint(
            course_name=course_name_clean,
            month_key=month_key_clean,
            players=players,
        )
        image_hash = st.session_state.get("ocr_uploaded_image_hash")
        registry = _load_processed_scorecards_registry()
        duplicate_kind, duplicate_item = _find_duplicate_scorecard(registry, fingerprint, image_hash)
        if duplicate_kind:
            st.error(
                "Duplicate scorecard detected and skipped "
                f"({duplicate_kind} already processed for "
                f"{duplicate_item.get('course', '-')}, {duplicate_item.get('month_key', '-')})."
            )
            return

        goam_scores[month_key_clean] = {
            "course": course_name_clean,
            "players": players,
        }
        save_json("data/goam_scores.json", goam_scores)

        _register_processed_scorecard(
            registry=registry,
            fingerprint=fingerprint,
            image_hash=image_hash,
            month_key=month_key_clean,
            course_name=course_name_clean,
        )

        course_score_path, course_score_error = _build_course_score_from_template(
            course_name=course_name_clean,
            month_key=month_key_clean,
            player_rows=players,
        )

        st.success(f"Published OCR scorecard to data/goam_scores.json under {month_key_clean}.")
        if course_score_error:
            st.warning(course_score_error)
        elif course_score_path:
            st.success(f"Created course score file: {course_score_path}")


# ---------------------------------------------------------
# MAIN ENTRY POINT
# ---------------------------------------------------------
def run_scores_app(mode="leaderboards"):
    if mode == "leaderboards":
        show_leaderboards()

    elif mode == "scorecards":
        show_scorecards()

    elif mode == "scoreboard_poster":
        show_scoreboard_poster()

    elif mode == "scorecard_ocr":
        show_scorecard_ocr_reader()

    else:
        st.error("Invalid mode for run_scores_app()")
